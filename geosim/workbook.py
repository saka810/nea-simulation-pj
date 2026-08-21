"""結果一式を 1 つの Excel ファイル（.xlsx）にまとめる（依頼 2026-08-21）。

> 結果ファイルもそうですが、CSV の方が、逐次読み込む際に都合が良いですか？
> .xlsx ファイルでまとめた方が、例えば、グラフを作ってもらうのもお願い、あるいは
> テンプレ準備しておいてそれにあてこむだけ、というのが出来るかなと思いました。
> 場合によっては、ソフトで読み込んだりする用のデータと、体裁を整えたりできる用の
> ファイル（結果一式まとまったファイル）を分けても良いかもしれません。

## 役割を分ける

**分けるのが正解**というのがこのモジュールの立場。

| | CSV（`結果/**.csv`） | Excel（`結果/<室>_結果一式.xlsx`） |
|---|---|---|
| 誰が読むか | プログラム（`--redraw`・まとめ表・他ソフト） | 人（報告書・打ち合わせ） |
| 良いところ | 1 行ずつ読める・差分が見える・文字コードだけで開ける | 1 枚で全部見える・**グラフが付く**・体裁を作れる |
| 都合が悪いところ | 枚数が多い・書式が持てない | 逐次読みに向かない・ライブラリが要る |

だから**CSV は今のまま残し、その上に Excel を 1 枚かぶせる**。
Excel 側は「CSV を読んで並べ直すだけ」なので、計算はやり直さない。

    python workbook.py <プロジェクト>                  … 結果一式.xlsx を作る
    python workbook.py <プロジェクト> --template 雛形.xlsx

## 中身

| シート | 内容 | グラフ |
|---|---|---|
| `概要` | 計算条件と主要な結果（代表値） | — |
| `残響時間` | 全受音点 ＋ 平均 ＋ ばらつき ＋ 理論値 | 折れ線 |
| `明瞭度` | C50 / C80 / D50 / Ts | 折れ線 |
| `音圧レベル` | 帯域別 Lp と**自由音場（逆二乗）との差** | 折れ線 |
| `STI` | STI と帯域別 MTI | 棒 |
| `吸音率と理論値` | 材料別の吸音率 → 平均吸音率 → 残響時間理論値 | 折れ線 |
| `材料条件表` | レイヤー名と吸音材の対応（入力の記録） | — |

## テンプレートに流し込む

`--template 雛形.xlsx` を渡すと、**その雛形を開いて同じ名前のシートに値だけ書く**。
雛形側で作ったグラフ・書式・ロゴはそのまま残るので、
体裁を毎回作り直さずに済む（雛形のセル参照が動かないよう、
**書き出す位置は必ず A1 から**にしてある）。

雛形に無いシート名は新しく作る。**雛形のシートを消すことはしない。**
"""

import os

import numpy as np

import project as pj
import summary as sm
import table as tb

WORKBOOK_FILE = "結果一式.xlsx"
TEMPLATE_FILE = "テンプレート.xlsx"

# シート名（テンプレート側と突き合わせるキーになるので**変えると雛形が外れる**）
SHEET_OVERVIEW = "概要"
SHEET_REVERBERATION = "残響時間"
SHEET_CLARITY = "明瞭度"
SHEET_LEVEL = "音圧レベル"
SHEET_STI = "STI"
SHEET_ROOM = "吸音率と理論値"
SHEET_CONDITION = "材料条件表"

HEADER_FILL = "FFEFF3F8"
TITLE_FONT_SIZE = 11


def path(project):
    """結果一式（xlsx）のパス。`結果/` 直下（受音点に依らないので）。"""
    return os.path.join(project.path(pj.RESULT_DIR),
                        project.prefixed(WORKBOOK_FILE))


def template_path(project, template=None):
    """使う雛形のパス。指定が無ければプロジェクト直下の `テンプレート.xlsx`。"""
    if template:
        return template if os.path.exists(template) else None
    candidate = project.path(TEMPLATE_FILE)
    return candidate if os.path.exists(candidate) else None


# ------------------------------------------------------------------------------
# CSV を読んで「シート 1 枚ぶんの表」にする
# ------------------------------------------------------------------------------

def _read_rows(path_):
    """CSV をそのまま行の list で読む。無ければ None。"""
    import csv
    if not path_ or not os.path.exists(path_):
        return None
    with open(path_, encoding="utf-8-sig", newline="") as f:
        rows = [row for row in csv.reader(f)
                if row and any(c.strip() for c in row)]
    return rows or None


def _numeric(rows, skip_columns):
    """文字列の表を、数値になるセルだけ数値に直す（Excel でグラフにできるように）。"""
    out = []
    for i, row in enumerate(rows):
        cells = []
        for j, value in enumerate(row):
            text = (value or "").strip()
            if i == 0 or j < skip_columns or not text:
                cells.append(text)
                continue
            try:
                cells.append(float(text))
            except ValueError:
                cells.append(text)
        out.append(cells)
    return out


def _summary_table(project, filename, skip_columns):
    rows = _read_rows(os.path.join(project.path(pj.RESULT_DIR),
                                   project.prefixed(filename)))
    return None if rows is None else _numeric(rows, skip_columns)


def _overview(project):
    """条件と主要な結果を「項目・値」の 2 列で並べる。"""
    import absorption as ab

    kind = {"normal": "垂直入射", "random": "残響室法"}.get(project.absorption_kind,
                                                            "未指定")
    rows = [["項目", "値"],
            ["対象室・条件名", project.name],
            ["モデル（DXF）", project.dxf],
            ["吸音率表", f"{project.absorption_csv}（{kind}）"],
            ["周波数バンド", f"{project.band_number}（"
                             f"{ab.octave_bands(project.band_number)[0]:.0f}〜"
                             f"{ab.octave_bands(project.band_number)[-1]:.0f} Hz）"],
            ["音線の本数", project.rays],
            ["最大反射回数", project.nref],
            ["受音球の半径 [m]", project.radius],
            ["応答の長さ [s]", project.max_time],
            ["温度 [℃] / 湿度 [%] / 気圧 [kPa]",
             f"{project.temperature} / {project.humidity} / {project.pressure}"],
            ["音源パワーレベル PWL [dB]",
             "未入力（音圧レベルは相対値）" if project.source_power_db is None
             else str(project.source_power_db)],
            ["背景騒音 [dB]",
             "未入力（STI は騒音なしで計算）" if project.noise_level_db is None
             else str(project.noise_level_db)],
            [None, None]]

    room = pj.read_room_csv(project.existing_result_path("room"))
    if room is not None:
        total = room["surface"]["total_area"]
        rows += [["総表面積 [m2]", round(total, 3)]]
    if project.volume:
        rows += [["室容積 [m3]（指定値）", project.volume]]

    receivers = sm.receiver_folders(project)
    rows += [["受音点の数", len(receivers)], [None, None], ["主要な結果", ""]]

    # 代表値は「500 Hz に最も近いバンド」で拾う（中域の代表として実務で使う）
    def representative(filename, item, skip):
        rows_ = _summary_table(project, filename, skip)
        if rows_ is None:
            return None
        header = rows_[0]
        try:
            frequencies = np.array([float(v) for v in header[skip:]])
        except ValueError:
            return None
        column = skip + int(np.argmin(np.abs(frequencies - 500.0)))
        for row in rows_[1:]:
            if row[0] == "平均" and row[1] == item:
                return row[column], frequencies[column - skip]
        return None

    for filename, item, skip, label in (
            (sm.REVERBERATION_FILE, "T30_s", 2, "T30（平均）[s]"),
            (sm.REVERBERATION_FILE, "EDT_s", 2, "EDT（平均）[s]"),
            (sm.CLARITY_FILE, "C50_db", 2, "C50（平均）[dB]"),
            (sm.LEVEL_FILE, "Lp_dB", 3, "音圧レベル（平均）[dB]")):
        found = representative(filename, item, skip)
        if found is not None:
            rows.append([f"{label}（{found[1]:.0f} Hz）", found[0]])

    sti = _summary_table(project, sm.STI_FILE, 3)
    if sti is not None:
        for row in sti[1:]:
            if row[0] == "平均" and row[1] in ("STI", "評価"):
                rows.append([f"STI（平均）" if row[1] == "STI" else "STI の評価",
                             row[2]])
    return rows


def sheets(project, verbose=True):
    """書き出すシートを [(シート名, 行の list, グラフの種類, 見出しの列数), …] で返す。

    グラフの種類は 'line' / 'bar' / None。見出しの列数は
    「凡例に使う左側の列がいくつあるか」（`受音点,項目` なら 2）。
    """
    import condition_table as ct

    result = [(SHEET_OVERVIEW, _overview(project), None, 1)]
    for name, filename, chart, skip in (
            (SHEET_REVERBERATION, sm.REVERBERATION_FILE, "line", 2),
            (SHEET_CLARITY, sm.CLARITY_FILE, "line", 2),
            (SHEET_LEVEL, sm.LEVEL_FILE, "line", 3),
            (SHEET_STI, sm.STI_FILE, "bar", 3),
            (SHEET_ROOM, pj.RESULT_FILES["room"], "line", 3)):
        rows = _summary_table(project, filename, skip)
        if rows is not None:
            result.append((name, rows, chart, skip))
        elif verbose:
            print(f"[Excel] {filename} が無いので『{name}』は飛ばします")

    condition = _read_rows(ct.path(project))
    if condition is not None:
        # `#` で始まる覚え書きの行は落とす（Excel では邪魔になる）
        condition = [row for row in condition if not row[0].startswith("#")]
        result.append((SHEET_CONDITION, _numeric(condition, 3), None, 3))
    return result


# ------------------------------------------------------------------------------
# 書き出す
# ------------------------------------------------------------------------------

def write(project, template=None, verbose=True):
    """結果一式の xlsx を書き出す。書いたパスを返す。

    `template` に雛形を渡すと**その体裁のまま**同じ名前のシートへ値を書く。
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise ImportError("Excel 出力には openpyxl が要ります。"
                          "`pip install -r requirements.txt` を実行してください")

    blocks = sheets(project, verbose=verbose)
    source = template_path(project, template)
    if source:
        book = load_workbook(source)
        if verbose:
            print(f"[Excel] 雛形を使います: {source}")
    else:
        book = Workbook()
        book.remove(book.active)

    for name, rows, chart, skip in blocks:
        sheet = book[name] if name in book.sheetnames else book.create_sheet(name)
        _fill(sheet, rows)
        if source is None:
            _decorate(sheet, rows, skip)
            if chart and len(rows) > 1:
                _add_chart(sheet, name, rows, chart, skip)

    out = path(project)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    book.save(out)
    if verbose:
        print(f"[Excel] 結果一式（{len(blocks)} シート）: {out}")
    return out


def _fill(sheet, rows):
    """A1 から値を書く。**雛形のセル参照がずれないよう位置は固定。**"""
    for i, row in enumerate(rows, start=1):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)


def _decorate(sheet, rows, skip):
    """見出しを太字にし、列幅とウィンドウ枠を整える（雛形が無いときだけ）。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=TITLE_FONT_SIZE)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    widths = {}
    for row in rows:
        for j, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            widths[j] = max(widths.get(j, 8), min(len(text) + 2, 42))
    for j, width in widths.items():
        sheet.column_dimensions[get_column_letter(j)].width = width
    # 見出しと左側の項目名を固定して、右へスクロールしても読めるようにする
    sheet.freeze_panes = sheet.cell(row=2, column=skip + 1)

    for row in sheet.iter_rows(min_row=2, min_col=skip + 1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"


def _add_chart(sheet, name, rows, kind, skip):
    """行を系列、周波数を横軸にしたグラフを右側に置く。

    表と同じ向き（**周波数は横**）なので `from_rows=True` でそのまま系列になる。
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter

    columns = len(rows[0])
    if columns <= skip + 1:
        return
    chart = LineChart() if kind == "line" else BarChart()
    chart.title = name
    chart.y_axis.title = name
    chart.x_axis.title = "周波数 [Hz]"
    chart.height, chart.width = 9.0, 20.0

    data = Reference(sheet, min_col=skip + 1, max_col=columns,
                     min_row=2, max_row=len(rows))
    categories = Reference(sheet, min_col=skip + 1, max_col=columns,
                           min_row=1, max_row=1)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(categories)
    # 系列名は「受音点＋項目」を並べたもの（表の左側の列をつなげる）
    for series, row in zip(chart.series, rows[1:]):
        label = " ".join(str(v) for v in row[:skip] if v)
        series.tx = None
        try:
            from openpyxl.chart.series import SeriesLabel
            from openpyxl.chart.data_source import StrRef
            series.tx = SeriesLabel(v=label)
        except Exception:      # 系列名が付かなくてもグラフ自体は出る
            pass
    sheet.add_chart(chart, f"{get_column_letter(columns + 2)}2")
    return chart


def main():
    import argparse

    p = argparse.ArgumentParser(
        description="結果一式を 1 つの Excel ファイルにまとめる（計算はやり直さない）")
    p.add_argument("folder", help="プロジェクトフォルダ")
    p.add_argument("--template", help="雛形の xlsx（体裁とグラフを残したまま値を入れる）")
    a = p.parse_args()

    project = pj.Project.load(a.folder)
    print(write(project, template=a.template))


if __name__ == "__main__":
    main()
