"""条件表 ― 「レイヤー名 → 材料番号」の対応と吸音率を 1 つの Excel で持つ。

## かたち

    プロジェクトフォルダ/条件表.xlsx

      シート「吸音率」    ← PJ で使う材料の一覧。**1000 材料ぶんの枠**
        番号,材料名,63,125,…,8000,種類,備考
         1 ,スチールドア,0.15,0.29,…,残響室法,
        （番号は `=ROW()-1`。**行の位置がそのまま材料番号**）

      シート「現状」      ← 条件 1 つ ＝ シート 1 枚。**シート名が条件名**。30 行の枠
        番号,区分,レイヤー名,材料番号,安全率,材料名（参考）,面数,面積_m2
         1 ,レイヤ,01__研修室_壁_扉,1,,=IFERROR(VLOOKUP(...)),2,3.3

      シート「吸音追加案」 ← シートを複製して番号を書き換えるだけ

### ★入力は「材料番号」と「安全率」だけ

材料名で書くと**全角・半角のミスタイプを誘発する**（ユーザー指摘）。
材料名の列は `VLOOKUP` で吸音率シートを引く**参考**で、読むときは見ない。

**安全率**は「吸音率を見過ぎないように間引く」ための係数（例 0.8）。
**未入力なら掛けない。**カタログ値（＝吸音率シートの値）に掛けてから
垂直入射へ変換するので、掛ける順番も実務の考え方と同じ。

### ★吸音率は条件シートに載せない

参考として α を並べるのは**ミスリードを誘発する**のでやめた（ユーザー指摘）。
吸音率は「吸音率」シートにだけ置く。そのぶん**PJ 固有の吸音データを
この 1 ファイルに閉じ込められる**（`absorption.csv` を別に持ち回らなくてよい）。

### 表示の約束（ユーザー指定 2026-08-21）

| 列 | 書式 |
|---|---|
| 吸音率 α・安全率 | 小数 2 桁（`0.00`） |
| 面積 | 小数 1 桁（`0.0`） |
| 面数・番号 | 整数 |

### 「区分」列について

いまある値は 3 つ。

| 区分 | 意味 | 入力か記録か |
|---|---|---|
| `レイヤ` | DXF のレイヤに材料番号を割り当てる | **入力** |
| `面ごとの指定` | `face_editor` で面を選んで貼った材料 | 記録 |
| `（モデルに無し）` | 前は書いてあったが、いまのモデルに無いレイヤ | 記録 |

**将来ありうるもの**（まだ実装していない）:

- `追加吸音` … 客席・人・什器のように**形を作らず面積だけで足したい吸音**
  （等価吸音面積 A [m²] を直接足す。実務でよく使う）
- `グループ` … 複数のレイヤをまとめて 1 つの材料にする行

★区分が要らないと判断したら、`面ごとの指定` と `（モデルに無し）` を別シートに
移して列を落とせる。**すぐには消さない**（記録の置き場が無くなるため）。

## 更新の約束（★ここを崩さないこと）

**既存の xlsx は「その場で直す」。書き直さない。**
利用者が作った体裁（列幅・書式・番号の式・列の並び・増やした列）を壊さないため、
`update()` は見出しの文字から列を探し、**面数・面積・参考の式だけ**を書き換える。
材料番号と安全率は**絶対に上書きしない**。モデルから消えたレイヤは区分を
`（モデルに無し）` にして残す（前の設定を見返せるように）。
"""

import csv
import os

import numpy as np

import absorption as ab

# 既定のファイル名
CONDITION_BOOK = "条件表.xlsx"
CONDITION_FILE = "材料条件表.csv"      # 昔の形（CSV）。読むためだけに残す

# 吸音率のシート名（**条件のシートではない**ので一括計算からは外す）
ABSORPTION_SHEET = "吸音率"
# 条件として扱わないシート名（説明や凡例を置けるように）
RESERVED_SHEETS = {ABSORPTION_SHEET, "説明", "凡例", "メモ", "readme", "README"}

# 区分の名前
SECTION_LAYER = "レイヤ"
SECTION_FACE = "面ごとの指定"
SECTION_GONE = "（モデルに無し）"

# 枠の数（ユーザー指定 2026-08-21）
MATERIAL_SLOTS = 1000      # 吸音率シートの材料の枠
LAYER_SLOTS = 30           # 条件シートのレイヤの枠

# 書式（ユーザー指定：吸音率は小数 2 桁、面積は小数 1 桁）
ALPHA_FORMAT = "0.00"
AREA_FORMAT = "0.0"
COUNT_FORMAT = "0"

# 条件シートの列。**見出しの文字で探す**ので、並び順は自由に変えられる。
# `番号` は行の通し番号（飾り）。読み書きしない
COLUMN_INDEX = "番号"
COLUMN_SECTION = "区分"
COLUMN_LAYER = "レイヤー名"
COLUMN_NUMBER = "材料番号"
COLUMN_FACTOR = "安全率"
COLUMN_NAME = "材料名（参考）"
COLUMN_COUNT = "面数"
COLUMN_AREA = "面積_m2"

HEADER = [COLUMN_INDEX, COLUMN_SECTION, COLUMN_LAYER, COLUMN_NUMBER,
          COLUMN_FACTOR, COLUMN_NAME, COLUMN_COUNT, COLUMN_AREA]
# 見出しが無くても困らない列（古い表との互換）
OPTIONAL_COLUMNS = {COLUMN_INDEX, COLUMN_FACTOR, COLUMN_NAME}

# 昔の CSV の見出し（3 列目が材料名だった）
LEGACY_HEADER = ["区分", "レイヤー名", "材料名", "面数", "面積_m2"]

# 吸音率シートの見出し
ABSORPTION_HEADER_HEAD = ["番号", "材料名"]
ABSORPTION_HEADER_TAIL = ["種類", "備考"]

# 吸音率の種類（**リストから選ぶ**。ユーザー要望 2026-08-21）。
# 取り違えると吸音を大きく誤るので、日本語で選べるようにしてある
KIND_LABELS = {"残響室法": ab.KIND_RANDOM, "垂直入射": ab.KIND_NORMAL,
               "random": ab.KIND_RANDOM, "normal": ab.KIND_NORMAL}
KIND_CHOICES = ("残響室法", "垂直入射")
KIND_TEXT = {ab.KIND_RANDOM: "残響室法", ab.KIND_NORMAL: "垂直入射"}

# 既定で作る条件シートの名前
FIRST_SHEET = "現状"


def kind_of(text, default=None):
    """種類の文字（`残響室法` / `normal` など）を `absorption` の種類に直す。"""
    return KIND_LABELS.get(str(text or "").strip(), default)


# ------------------------------------------------------------------------------
# 場所とシート
# ------------------------------------------------------------------------------

def path(project):
    """使う条件表のパス（`project.condition_csv`、無ければ既定名）。"""
    return project.condition_path


def sheet_of(project):
    """使う条件シートの名前。指定が無ければ最初の条件シート。"""
    if project.condition_sheet:
        return project.condition_sheet
    found = sheets(path(project))
    return found[0] if found else None


def exists(project):
    return os.path.exists(path(project))


def is_book(file_name):
    return bool(file_name) and str(file_name).lower().endswith((".xlsx", ".xlsm"))


def sheets(file_name):
    """条件表の中の**条件シートの名前**を並び順に返す（吸音率などは除く）。

    CSV なら [None]（シートの概念が無い）。読めなければ []。
    """
    if not file_name or not os.path.exists(file_name):
        return []
    if not is_book(file_name):
        return [None]
    try:
        from openpyxl import load_workbook
        book = load_workbook(file_name, read_only=True)
    except Exception as error:
        print(f"[条件表] {file_name} を開けませんでした: "
              f"{type(error).__name__}: {error}")
        return []
    try:
        return [name for name in book.sheetnames if name not in RESERVED_SHEETS]
    finally:
        book.close()


def discover(folder, verbose=False):
    """フォルダの中の条件を全部返す [(パス, シート名), …]。一括計算の入力。

    `条件表.xlsx` の**条件シートを 1 件ずつ**に展開する。
    昔の CSV（`材料条件表.csv` の形）も拾う（シート名は None）。
    """
    found, books, legacy = [], [], []
    for name in sorted(os.listdir(folder)):
        candidate = os.path.join(folder, name)
        if is_book(name) and not name.startswith("~$"):
            if _looks_like_condition_book(candidate):
                books.append(candidate)
                found.extend((candidate, sheet) for sheet in sheets(candidate))
        elif name.lower().endswith(".csv") and is_condition_table(candidate):
            # 既定名の CSV（`材料条件表.csv`）は**xlsx があれば無視する**。
            # 移行後に同じ内容が 2 条件として並ぶのを防ぐ（2026-08-21）
            stem = os.path.splitext(name)[0]
            if stem in ("条件表", "材料条件表"):
                legacy.append(candidate)
            else:
                found.append((candidate, None))
    if not books:
        found.extend((candidate, None) for candidate in legacy)
    elif legacy and verbose:
        print(f"[条件表] 昔の CSV は使いません（xlsx があるので）: "
              + " / ".join(os.path.basename(f) for f in legacy))
    if verbose:
        print(f"[条件表] {folder} に条件が {len(found)} 件: "
              + " / ".join(label_of(p, s) for p, s in found))
    return found


def label_of(file_name, sheet=None):
    """条件の名前。**シートがあればシート名**、無ければファイル名（拡張子なし）。"""
    if sheet:
        return str(sheet)
    return os.path.splitext(os.path.basename(str(file_name or "")))[0]


def _looks_like_condition_book(file_name):
    """その xlsx が条件表かどうか（条件シートの見出しで判別する）。"""
    try:
        from openpyxl import load_workbook
        book = load_workbook(file_name, read_only=True)
    except Exception:
        return False
    try:
        for name in book.sheetnames:
            row = next(book[name].iter_rows(max_row=1, values_only=True), None)
            if row and _columns_of([_text(c) for c in row]):
                return True
    except Exception:
        return False
    finally:
        book.close()
    return False


def is_condition_table(file_name):
    """その CSV が条件表かどうか。見出し行で判別する（昔の形も通る）。"""
    if not os.path.isfile(file_name):
        return False
    try:
        with open(file_name, encoding="utf-8-sig", newline="") as f:
            for line in f:
                if not line.strip() or line.startswith("#"):
                    continue
                cells = [c.strip() for c in line.split(",")]
                return bool(_columns_of(cells))
    except (OSError, UnicodeDecodeError):
        return False
    return False


def _columns_of(header_cells):
    """見出しの並びから {列の名前: 0 始まりの位置} を作る。条件表でなければ {}。

    ★**見出しの文字で探す**のが肝。利用者が列を増やしたり並べ替えたりしても
    読めるようにするため（`番号` の列を足す、など）。
    """
    found = {}
    for index, cell in enumerate(header_cells):
        text = str(cell or "").strip()
        if not text:
            continue
        if text in HEADER:
            found.setdefault(text, index)
        elif text == "材料名":          # 昔の CSV（3 列目が材料名）
            found.setdefault(COLUMN_NUMBER, index)
        elif text.startswith("材料名"):  # 「材料名（参考）」の表記揺れ
            found.setdefault(COLUMN_NAME, index)
    # レイヤー名と（材料番号 か 材料名）がそろっていれば条件表とみなす
    if COLUMN_LAYER in found and COLUMN_NUMBER in found:
        return found
    return {}


# ------------------------------------------------------------------------------
# 読む
# ------------------------------------------------------------------------------

def read(file_name, sheet=None):
    """条件表を読む。戻り値 (割り当て, 記録).

    割り当て … {レイヤー名: 材料番号}（**番号の列だけを見る**。空の行は入れない）
    記録     … [(区分, レイヤー名, 材料番号, 面数, 面積, 安全率), …]
    """
    rows = _rows_of(file_name, sheet)
    if not rows:
        return {}, []
    columns = _columns_of(rows[0])
    if not columns:
        return {}, []

    def cell(row, name):
        index = columns.get(name)
        return "" if index is None or index >= len(row) else row[index].strip()

    assignment, records = {}, []
    for row in rows[1:]:
        layer = cell(row, COLUMN_LAYER)
        if not layer:
            continue          # 空の枠（30 行ぶん用意してあるので普通にある）
        section = cell(row, COLUMN_SECTION) or SECTION_LAYER
        number = cell(row, COLUMN_NUMBER)
        records.append((section, layer, number, cell(row, COLUMN_COUNT),
                        cell(row, COLUMN_AREA), cell(row, COLUMN_FACTOR)))
        if section == SECTION_LAYER and number:
            assignment[layer] = number
    return assignment, records


def safety_factors(file_name, sheet=None):
    """{レイヤー名: 安全率} を読む。**未入力・1.0・読めない値は入れない**。

    「吸音率を見過ぎないように間引く」ための係数（ユーザー要望 2026-08-21）。
    """
    _, records = read(file_name, sheet)
    factors = {}
    for section, layer, _number, _count, _area, factor in records:
        if section != SECTION_LAYER or not factor:
            continue
        try:
            value = float(factor)
        except ValueError:
            print(f"[条件表] 安全率が数値になっていません（{layer}: {factor!r}）。"
                  f"掛けずに進めます")
            continue
        if value <= 0.0:
            print(f"[条件表] 安全率 {value} は使えません（{layer}）。掛けずに進めます")
            continue
        if not np.isclose(value, 1.0):
            factors[layer] = value
    return factors


def _rows_of(file_name, sheet=None):
    if not file_name or not os.path.exists(file_name):
        return []
    return (_book_rows(file_name, sheet) if is_book(file_name)
            else _csv_rows(file_name))


def _csv_rows(file_name):
    with open(file_name, encoding="utf-8-sig", newline="") as f:
        return [list(row) for row in csv.reader(f)
                if row and row[0].strip() and not row[0].startswith("#")]


def _book_rows(file_name, sheet=None):
    """xlsx の 1 シートを文字列の表で読む。式はそのまま（読む列は式でないので）。"""
    from openpyxl import load_workbook

    book = load_workbook(file_name, read_only=True)
    try:
        name = sheet or next((n for n in book.sheetnames
                              if n not in RESERVED_SHEETS), None)
        if name is None or name not in book.sheetnames:
            return []
        return [[_text(v) for v in values]
                for values in book[name].iter_rows(values_only=True)
                if values is not None and any(v is not None and str(v).strip()
                                              for v in values)]
    finally:
        book.close()


def _text(value):
    """セルの値を文字列にする。**整数はそのまま**（`1.0` にしない）。式は空扱い。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    # 配列数式のオブジェクトや式の文字列は「入力値」ではないので空にする
    if text.startswith("=") or "openpyxl" in text:
        return ""
    return text


def assignment_for(project, verbose=True):
    """条件表から {レイヤー名: 材料番号} を返す。無ければ project.json の指定。"""
    if not exists(project):
        if project.condition_csv:
            print(f"[条件表] 指定された条件表が見つかりません: "
                  f"{project.condition_csv}（レイヤー名で吸音材を引きます）")
        return project.assignment
    sheet = sheet_of(project)
    assignment, _ = read(path(project), sheet)
    if verbose:
        where = os.path.basename(path(project))
        if sheet:
            where += f" のシート『{sheet}』"
        print(f"[条件表] {where} から {len(assignment)} レイヤの割り当てを読みました")
    if project.assignment and verbose:
        print("[条件表] 注意: project.json 側の割り当ては使いません（条件表を優先）")
    return assignment or None


def factors_for(project, verbose=True):
    """条件表から {レイヤー名: 安全率} を返す。無ければ空。"""
    if not exists(project):
        return {}
    factors = safety_factors(path(project), sheet_of(project))
    if factors and verbose:
        print("[条件表] 安全率: "
              + " / ".join(f"{k} ×{v:g}" for k, v in sorted(factors.items())))
    return factors


def library_from_book(file_name, kind=None, verbose=True):
    """条件表の「吸音率」シートから材料一覧を作る。無ければ None。

    ★**番号は行の位置から決める**（`番号` 列が `=ROW()-1` の式で、
    Excel で一度も開いていないと計算結果が入っていないため）。
    列に数値が入っていればそれを優先する。
    """
    if not is_book(file_name) or not os.path.exists(file_name):
        return None
    try:
        from openpyxl import load_workbook
        book = load_workbook(file_name, read_only=True, data_only=True)
    except Exception as error:
        print(f"[条件表] {file_name} を開けませんでした: "
              f"{type(error).__name__}: {error}")
        return None
    try:
        if ABSORPTION_SHEET not in book.sheetnames:
            return None
        rows = [list(v) for v in book[ABSORPTION_SHEET].iter_rows(values_only=True)]
    finally:
        book.close()

    library = ab.MaterialLibrary()
    unknown_kind = 0
    for offset, row in enumerate(rows[1:] if rows else [], start=1):
        cells = [_text(v) for v in row]
        if len(cells) < 3 or not cells[1]:
            continue
        # 番号は列の値 → 無ければ行の位置（`=ROW()-1` と同じ意味）
        number = cells[0] if cells[0].isdigit() else str(offset)
        name = cells[1]
        values, material_kind = [], None
        for cell in cells[2:]:
            resolved = kind_of(cell)
            if resolved is not None:
                material_kind = resolved
                break
            try:
                values.append(float(cell))
            except ValueError:
                break
        if len(values) < 3:
            continue
        if material_kind is None:
            material_kind = kind or ab.KIND_NORMAL
            unknown_kind += 1
        library.add(name, values, kind=material_kind, overwrite=True)
        library.aliases[number] = name
    if not len(library):
        return None
    if verbose:
        print(f"[条件表] 「{ABSORPTION_SHEET}」シートから {len(library)} 材料"
              f"（{os.path.basename(file_name)}）")
        if unknown_kind:
            print(f"[条件表] 注意: 種類（残響室法／垂直入射）が空の材料が "
                  f"{unknown_kind} 件あります。"
                  f"{'指定の' if kind else '既定の'}"
                  f"{KIND_TEXT.get(kind or ab.KIND_NORMAL)}として扱います。"
                  f"★取り違えると吸音を大きく誤るので確認してください")
    return library


def absorption_table(library, assignment, factors=None, band_number=None,
                     warn=True):
    """吸音率テーブルを作り、**安全率を掛ける**。

    ★安全率は**カタログ値（吸音率シートの値）に掛けてから**垂直入射へ変換する。
    「カタログの 8 割で見ておく」という実務の考え方と順番を合わせるため
    （変換後に掛けると意味が変わる）。
    """
    table = library.absorption_table(assignment, band_number=band_number,
                                     warn=warn)
    if not factors:
        return table
    mapping = (assignment.mapping if hasattr(assignment, "mapping")
               else dict(assignment or {}))
    bands = band_number or len(next(iter(table.values()), [0] * 8))
    for layer, factor in factors.items():
        material = library.get(mapping.get(layer, layer))
        if material is None:
            if warn:
                print(f"[条件表] 安全率を掛ける材料が引けません（{layer}）")
            continue
        scaled = ab.Material(material.name,
                             np.asarray(material.resample(bands)) * factor,
                             material.kind, material.note)
        table[layer] = scaled.normal_incidence(warn=warn)
    return table


# ------------------------------------------------------------------------------
# 作る・更新する
# ------------------------------------------------------------------------------

def resolve_material(layer, library, assignment):
    """そのレイヤに使われる材料の**キー**（番号か名前）を返す。引けなければ ""。

    引く順番は `read_dxffile._resolve_absorption` と同じ。
        ① 条件表／project.json の割り当て
        ② レイヤー名そのもの（材料名または別名）
        ③ レイヤー名の先頭の番号（元コードの材料 ID）
    """
    if assignment:
        name = (assignment.mapping if hasattr(assignment, "mapping")
                else assignment).get(layer)
        if name:
            return name
    if library is None:
        return ""
    if layer in library:
        return layer
    import read_dxffile as rd
    number = rd.layer_number(layer)
    if number is not None and number in library:
        return number
    return ""


def material_name(key, library):
    """材料番号（または名前）から材料名を引く。引けなければ ""。"""
    if not key or library is None:
        return ""
    if key in library.aliases:
        return library.aliases[key]
    return key if key in library.materials else ""


def create(project, model, library=None, sheet=FIRST_SHEET, verbose=True):
    """**DXF のレイヤから条件表（xlsx）を新しく作る。**

    設定画面の「条件表を作成」から呼ぶ。「吸音率」シート（1000 材料の枠）と
    条件シート（30 レイヤの枠）を作る。**既にあるファイルは書き直さず更新する**。
    """
    file_name = path(project)
    if not is_book(file_name):
        # 既定は xlsx。CSV を指定されている場合はその隣に xlsx を作る
        file_name = os.path.join(os.path.dirname(file_name) or project.folder,
                                 CONDITION_BOOK)
    if os.path.exists(file_name):
        return update(project, model, library, verbose=verbose)
    _new_book(file_name, project, model, library, sheet)
    if verbose:
        print(f"[条件表] 作りました（「{ABSORPTION_SHEET}」{MATERIAL_SLOTS} 枠 ＋ "
              f"「{sheet}」{LAYER_SLOTS} 枠）: {file_name}")
    return file_name


def update(project, model, library=None, assignment=None, verbose=True):
    """条件表を更新する。**xlsx は「その場で直す」（書き直さない）。**

    ★利用者が作った体裁（列幅・書式・番号の式・増やした列）を壊さないため、
    見出しの文字から列を探し、**面数・面積・参考の式だけ**を書き換える。
    材料番号と安全率は絶対に上書きしない。
    """
    file_name = path(project)
    if not is_book(file_name):
        previous, _ = read(file_name)
        if assignment is None:
            assignment = previous or project.assignment or {}
        _write_csv(file_name, project, model, library,
                   previous or _as_dict(assignment))
    elif not os.path.exists(file_name):
        return create(project, model, library, verbose=verbose)
    else:
        _update_book(file_name, project, model, library, verbose=verbose)
    if verbose:
        print(f"[条件表] 更新しました: {file_name}")
    return file_name


def _as_dict(assignment):
    if assignment is None:
        return {}
    return dict(assignment.mapping if hasattr(assignment, "mapping")
                else assignment)


def _face_records(project, model):
    """面ごとの割り当ての集計 {材料名: (面数, 面積)}。"""
    face_materials = project.face_materials_for(len(model.mesh))
    if not face_materials:
        return {}
    import reverberation as rv
    areas = rv.triangle_areas(model.mesh)
    result = {}
    for index, material in face_materials.items():
        if not material or index >= len(areas):
            continue
        count, area = result.get(material, (0, 0.0))
        result[material] = (count + 1, area + float(areas[index]))
    return result


# ---- 新しく作る --------------------------------------------------------------

def _new_book(file_name, project, model, library, sheet):
    from openpyxl import Workbook

    book = Workbook()
    book.remove(book.active)
    _build_absorption_sheet(book, project, library)
    _build_condition_sheet(book, sheet, project, model, library)
    book.save(file_name)
    return file_name


def _build_absorption_sheet(book, project, library):
    """「吸音率」シートを作る（1000 材料の枠・種類はリストから選ぶ）。"""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    frequencies = ab.octave_bands(project.band_number)
    sheet = book.create_sheet(ABSORPTION_SHEET)
    sheet.append(ABSORPTION_HEADER_HEAD + [f"{v:.0f}" for v in frequencies]
                 + ABSORPTION_HEADER_TAIL)

    first_alpha = len(ABSORPTION_HEADER_HEAD) + 1
    last_alpha = first_alpha + len(frequencies) - 1
    kind_column = last_alpha + 1

    # ★材料は**番号の行**に置く（番号 = 行の位置）。空きは枠として残す
    placed = {}
    for number, name, material in _materials_for(library):
        row = (int(number) + 1 if number.isdigit()
               and 1 <= int(number) <= MATERIAL_SLOTS else None)
        if row is None:
            continue
        placed[row] = (name, material)
    free = 2
    for number, name, material in _materials_for(library):
        if number.isdigit():
            continue
        while free in placed:
            free += 1
        placed[free] = (name, material)

    for row in range(2, MATERIAL_SLOTS + 2):
        sheet.cell(row=row, column=1, value=f"=ROW()-1")
        entry = placed.get(row)
        if entry is not None:
            name, material = entry
            sheet.cell(row=row, column=2, value=name)
            for i, value in enumerate(material.resample(project.band_number)):
                sheet.cell(row=row, column=first_alpha + i, value=float(value))
            sheet.cell(row=row, column=kind_column,
                       value=KIND_TEXT.get(material.kind, material.kind))
            if material.note:
                sheet.cell(row=row, column=kind_column + 1, value=material.note)
        for column in range(first_alpha, last_alpha + 1):
            sheet.cell(row=row, column=column).number_format = ALPHA_FORMAT

    # 種類はリストから選ぶ（取り違えると吸音を大きく誤るので）
    validation = DataValidation(
        type="list", allow_blank=True,
        formula1='"' + ",".join(KIND_CHOICES) + '"',
        showDropDown=False, errorTitle="吸音率の種類",
        error="残響室法（乱入射）か 垂直入射 を選んでください",
        promptTitle="吸音率の種類",
        prompt="カタログの出典を見て選ぶ。★取り違えると吸音を大きく誤ります")
    sheet.add_data_validation(validation)
    letter = get_column_letter(kind_column)
    validation.add(f"{letter}2:{letter}{MATERIAL_SLOTS + 1}")

    _decorate(sheet, [10, 34] + [8] * len(frequencies) + [12, 24])
    return sheet


def _build_condition_sheet(book, name, project, model, library, assignment=None):
    """条件シートを作る（30 レイヤの枠・材料名の式も 30 行ぶん入れる）。"""
    sheet = book.create_sheet(name)
    sheet.append(HEADER)
    columns = {label: i for i, label in enumerate(HEADER)}

    rows = _rows_for(project, model, library, _as_dict(assignment))
    for offset in range(max(LAYER_SLOTS, len(rows))):
        row = offset + 2
        record = rows[offset] if offset < len(rows) else None
        sheet.cell(row=row, column=columns[COLUMN_INDEX] + 1, value="=ROW()-1")
        if record is not None:
            section, layer, key, count, area = record
            sheet.cell(row=row, column=columns[COLUMN_SECTION] + 1, value=section)
            sheet.cell(row=row, column=columns[COLUMN_LAYER] + 1, value=layer)
            sheet.cell(row=row, column=columns[COLUMN_NUMBER] + 1,
                       value=_number_cell(key))
            sheet.cell(row=row, column=columns[COLUMN_COUNT] + 1, value=int(count))
            sheet.cell(row=row, column=columns[COLUMN_AREA] + 1,
                       value=round(float(area), 3))
        _prepare_row(sheet, row, columns)

    _decorate(sheet, [7, 14, 34, 12, 10, 34, 8, 12])
    return sheet


def _prepare_row(sheet, row, columns):
    """1 行ぶんの式と書式を整える（材料名の VLOOKUP・小数桁）。"""
    number_letter = _letter(columns[COLUMN_NUMBER] + 1)
    if COLUMN_NAME in columns:
        cell = sheet.cell(row=row, column=columns[COLUMN_NAME] + 1)
        section = (sheet.cell(row=row, column=columns[COLUMN_SECTION] + 1).value
                   if COLUMN_SECTION in columns else None)
        if section == SECTION_FACE:
            cell.value = None       # 面ごとの指定は番号を持たないので式を入れない
        else:
            # ★番号を入れたら名前がその場で出る。範囲は 1000 材料ぶん
            cell.value = (f'=IFERROR(VLOOKUP(${number_letter}{row},'
                          f'{ABSORPTION_SHEET}!$A$2:$B${MATERIAL_SLOTS + 1},'
                          f'2,FALSE),"")')
    if COLUMN_FACTOR in columns:
        factor = sheet.cell(row=row, column=columns[COLUMN_FACTOR] + 1)
        factor.number_format = ALPHA_FORMAT
    if COLUMN_AREA in columns:
        sheet.cell(row=row,
                   column=columns[COLUMN_AREA] + 1).number_format = AREA_FORMAT
    if COLUMN_COUNT in columns:
        sheet.cell(row=row,
                   column=columns[COLUMN_COUNT] + 1).number_format = COUNT_FORMAT


def _letter(index):
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


def _decorate(sheet, widths):
    from openpyxl.styles import Alignment, Font, PatternFill

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFEFF3F8")
        cell.alignment = Alignment(horizontal="center")
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[_letter(i)].width = width
    sheet.freeze_panes = "A2"


def _rows_for(project, model, library, assignment):
    """条件シート 1 枚ぶんの行（区分・レイヤー名・材料番号・面数・面積）。"""
    layer_areas = getattr(model, "layer_areas", {}) or {}
    layer_counts = getattr(model, "layer_counts", {}) or {}
    rows = []
    for layer in sorted(set(layer_counts) | set(layer_areas)):
        key = assignment.get(layer) or resolve_material(layer, library, assignment)
        rows.append((SECTION_LAYER, layer, key, layer_counts.get(layer, 0),
                     layer_areas.get(layer, 0.0)))
    for layer, key in sorted(assignment.items()):
        if layer not in layer_counts and layer not in layer_areas:
            rows.append((SECTION_GONE, layer, key, 0, 0.0))
    for material, (count, area) in sorted(_face_records(project, model).items()):
        rows.append((SECTION_FACE, material, "", count, area))
    return rows


def _materials_for(library):
    """吸音率シートに並べる材料 [(番号, 材料名, Material), …]。番号順。"""
    if library is None:
        return []
    numbers = {}
    for alias, target in library.aliases.items():
        if str(alias).strip().isdigit():
            numbers.setdefault(target, str(alias).strip())
    rows = [(numbers.get(name, ""), name, library.materials[name])
            for name in library.names()]
    rows.sort(key=lambda r: (r[0] == "", int(r[0]) if r[0].isdigit() else 0, r[1]))
    return rows


def _number_cell(key):
    """材料番号のセル。数字なら**数値として**入れる。空なら None。"""
    text = str(key or "").strip()
    if not text:
        return None
    return int(text) if text.isdigit() else text


# ---- その場で直す ------------------------------------------------------------

def _update_book(file_name, project, model, library, verbose=True):
    """既存の条件表を**その場で**直す。体裁と入力はそのまま残す。"""
    from openpyxl import load_workbook

    book = load_workbook(file_name)
    condition_sheets = [n for n in book.sheetnames if n not in RESERVED_SHEETS]
    if not condition_sheets:
        _build_condition_sheet(book, FIRST_SHEET, project, model, library)
        condition_sheets = [FIRST_SHEET]
    if ABSORPTION_SHEET not in book.sheetnames and library is not None:
        _build_absorption_sheet(book, project, library)
    elif ABSORPTION_SHEET in book.sheetnames:
        # 既にあるシートは値を触らず、**種類のリストと小数 2 桁だけ**整える
        _refresh_absorption_sheet(book[ABSORPTION_SHEET], project)

    layer_areas = getattr(model, "layer_areas", {}) or {}
    layer_counts = getattr(model, "layer_counts", {}) or {}
    faces = _face_records(project, model)

    for name in condition_sheets:
        sheet = book[name]
        columns = _columns_of([_text(c.value) for c in sheet[1]])
        if not columns:
            if verbose:
                print(f"[条件表] シート『{name}』は見出しが読めないので触りません")
            continue
        columns = _ensure_columns(sheet, columns)
        _refresh_sheet(sheet, columns, layer_counts, layer_areas, faces, library,
                       project, model)
    book.save(file_name)
    return file_name


def _refresh_absorption_sheet(sheet, project):
    """既にある「吸音率」シートに、種類のリストと吸音率の書式を足す。

    ★**値は触らない**（利用者が並べた材料をそのまま残す）。
    足りない仕掛け（ドロップダウン・小数 2 桁）だけ入れる。
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    header = [_text(c.value) for c in sheet[1]]
    kind_column = next((i + 1 for i, text in enumerate(header)
                        if text == "種類"), None)
    last = max(sheet.max_row, MATERIAL_SLOTS + 1)

    # 吸音率の列（番号・材料名より右で、種類より左）を小数 2 桁にする
    first_alpha = len(ABSORPTION_HEADER_HEAD) + 1
    last_alpha = (kind_column - 1) if kind_column else sheet.max_column
    for row in range(2, last + 1):
        for column in range(first_alpha, last_alpha + 1):
            sheet.cell(row=row, column=column).number_format = ALPHA_FORMAT

    if kind_column is None:
        return sheet
    # 既にある種類の書き方を日本語に揃える（意味は同じ。リストと食い違わないように）
    for row in range(2, min(last, sheet.max_row) + 1):
        cell = sheet.cell(row=row, column=kind_column)
        resolved = kind_of(cell.value)
        if resolved is not None and cell.value != KIND_TEXT[resolved]:
            cell.value = KIND_TEXT[resolved]
    letter = _letter(kind_column)
    target = f"{letter}2:{letter}{last}"
    for existing in sheet.data_validations.dataValidation:
        if existing.type == "list" and letter in str(existing.sqref):
            existing.sqref = target        # 範囲だけ広げる（1000 材料ぶん）
            return sheet
    validation = DataValidation(
        type="list", allow_blank=True,
        formula1='"' + ",".join(KIND_CHOICES) + '"',
        showDropDown=False, errorTitle="吸音率の種類",
        error="残響室法（乱入射）か 垂直入射 を選んでください",
        promptTitle="吸音率の種類",
        prompt="カタログの出典を見て選ぶ。★取り違えると吸音を大きく誤ります")
    sheet.add_data_validation(validation)
    validation.add(target)
    return sheet


def _ensure_columns(sheet, columns):
    """足りない列（安全率・材料名）を**右端に足す**。

    ★間に挿し込まない。既にある式（`VLOOKUP($D2,…)`）の参照がずれるため。
    """
    for label in (COLUMN_FACTOR, COLUMN_NAME):
        if label in columns:
            continue
        index = max(columns.values()) + 1
        while sheet.cell(row=1, column=index + 1).value not in (None, ""):
            index += 1
        cell = sheet.cell(row=1, column=index + 1, value=label)
        head = sheet.cell(row=1, column=1)
        cell.font, cell.fill, cell.alignment = head.font.copy(), \
            head.fill.copy(), head.alignment.copy()
        sheet.column_dimensions[_letter(index + 1)].width = (
            10 if label == COLUMN_FACTOR else 34)
        columns[label] = index
    return columns


def _refresh_sheet(sheet, columns, layer_counts, layer_areas, faces, library,
                   project, model):
    """1 枚の条件シートの面数・面積・式を書き直す（入力列は触らない）。"""
    section_column = columns.get(COLUMN_SECTION)
    layer_column = columns[COLUMN_LAYER]
    used, blanks = {}, []

    last = max(sheet.max_row, LAYER_SLOTS + 1)
    for row in range(2, last + 1):
        layer = _text(sheet.cell(row=row, column=layer_column + 1).value)
        if not layer:
            blanks.append(row)
            continue
        used[layer] = row

    def write(row, section, layer, count, area):
        if section_column is not None:
            sheet.cell(row=row, column=section_column + 1, value=section)
        sheet.cell(row=row, column=layer_column + 1, value=layer)
        if COLUMN_COUNT in columns:
            sheet.cell(row=row, column=columns[COLUMN_COUNT] + 1,
                       value=int(count))
        if COLUMN_AREA in columns:
            sheet.cell(row=row, column=columns[COLUMN_AREA] + 1,
                       value=round(float(area), 3))
        if COLUMN_INDEX in columns:
            cell = sheet.cell(row=row, column=columns[COLUMN_INDEX] + 1)
            if cell.value in (None, ""):
                cell.value = "=ROW()-1"
        _prepare_row(sheet, row, columns)

    # ① モデルにあるレイヤ
    for layer in sorted(set(layer_counts) | set(layer_areas)):
        row = used.pop(layer, None)
        if row is None:
            row = blanks.pop(0) if blanks else sheet.max_row + 1
            # 新しい行には自動で引けた番号を入れておく（利用者が直せる）
            key = resolve_material(layer, library, None)
            if key and COLUMN_NUMBER in columns:
                sheet.cell(row=row, column=columns[COLUMN_NUMBER] + 1,
                           value=_number_cell(key))
        write(row, SECTION_LAYER, layer, layer_counts.get(layer, 0),
              layer_areas.get(layer, 0.0))

    # ② 面ごとの指定（記録）
    for material, (count, area) in sorted(faces.items()):
        row = used.pop(material, None)
        if row is None:
            row = blanks.pop(0) if blanks else sheet.max_row + 1
        write(row, SECTION_FACE, material, count, area)

    # ③ 残ったもの＝いまのモデルに無い（前の設定として残す）
    for layer, row in used.items():
        write(row, SECTION_GONE, layer, 0, 0.0)

    # ④ 空の枠にも式と書式を入れておく（番号を書いたらすぐ名前が出るように）
    for row in blanks:
        if COLUMN_INDEX in columns:
            cell = sheet.cell(row=row, column=columns[COLUMN_INDEX] + 1)
            if cell.value in (None, ""):
                cell.value = "=ROW()-1"
        _prepare_row(sheet, row, columns)


def _write_csv(file_name, project, model, library, assignment):
    """昔の形（CSV）で更新する。**吸音率の列は書かない**（ミスリードを避ける）。"""
    with open(file_name, "w", encoding="utf-8-sig", newline="") as f:
        f.write("# 条件表 — レイヤー名と材料番号の対応\n")
        f.write(f"# 対象室: {project.room_label}\n")
        f.write(f"# モデル: {project.dxf}\n")
        f.write("# ★書き換えるのは「材料番号」と「安全率」の列だけ\n")
        writer = csv.writer(f)
        writer.writerow([COLUMN_SECTION, COLUMN_LAYER, COLUMN_NUMBER,
                         COLUMN_FACTOR, COLUMN_NAME, COLUMN_COUNT, COLUMN_AREA])
        previous = safety_factors(file_name)
        for section, layer, key, count, area in _rows_for(project, model, library,
                                                          assignment):
            writer.writerow([section, layer, key,
                             previous.get(layer, ""),
                             material_name(key, library), count, "%.1f" % area])
    return file_name


def main():
    import argparse

    p = argparse.ArgumentParser(description="条件表（xlsx）を作る／更新する")
    p.add_argument("folder", help="プロジェクトフォルダ")
    p.add_argument("--sheet", default=FIRST_SHEET, help="作る条件シートの名前")
    a = p.parse_args()

    import project as pj
    import run_project

    project = pj.Project.load(a.folder)
    library = run_project._library_for(project, verbose=True)
    model = run_project._model_for(project)
    print(create(project, model, library, sheet=a.sheet))


if __name__ == "__main__":
    main()
