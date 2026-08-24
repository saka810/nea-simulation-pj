# -*- coding: utf-8 -*-
"""細かい周波数で測った吸音率を**オクターブバンドに直す**。

★2026-08-24 ユーザー要望

> Excel データは一部ですが吸音率のデータです。3 条件あります。
> これを上手く補間して、63〜500 までの逆二乗を算出したい。

もらった表は 60・63・65・70・75… のように**不等間隔**で、抜けもある。
計算に使うのはオクターブバンド（63 / 125 / 250 / 500 …）なので、次の順で直す。

  ① 対数周波数で線形補間する（吸音率は周波数の対数に対してなだらかに動くので、
     等間隔でない測定値をつなぐならこれが素直）
  ② **バンドの幅で平均する**（1/1 オクターブなので f/√2 〜 f√2）。
     中心周波数の 1 点だけを読むより、山や谷を均せる
  ③ データの外側は**いちばん端の値を保つ**（外挿はしない）。
     どこから外挿かは `extrapolated` で返すので、出力に注意書きを添えられる

読むのは Excel でも CSV でもよい（`read_table`）。
"""
import io
import math
import os

import numpy as np

# バンドの幅を何点で均すか（対数周波数で等間隔）
SAMPLES = 33


def interpolate(frequencies, values, target):
    """`target` [Hz] の吸音率を**対数周波数の線形補間**で返す。"""
    frequencies = np.asarray(frequencies, dtype=float)
    values = np.asarray(values, dtype=float)
    order = np.argsort(frequencies)
    frequencies, values = frequencies[order], values[order]
    return float(np.interp(math.log10(float(target)),
                           np.log10(frequencies), values,
                           left=values[0], right=values[-1]))


def band_average(frequencies, values, centre, samples=SAMPLES):
    """1/1 オクターブバンド（`centre`）の平均吸音率。"""
    low, high = centre / math.sqrt(2.0), centre * math.sqrt(2.0)
    points = np.logspace(math.log10(low), math.log10(high), samples)
    return float(np.mean([interpolate(frequencies, values, f) for f in points]))


def to_bands(frequencies, values, centres, samples=SAMPLES):
    """オクターブバンドの値と、外挿になったバンドの一覧を返す。

    → (バンドごとの吸音率, 外挿したバンドの中心周波数)
    """
    frequencies = np.asarray(frequencies, dtype=float)
    lowest, highest = float(np.min(frequencies)), float(np.max(frequencies))
    bands, extrapolated = [], []
    for centre in centres:
        bands.append(band_average(frequencies, values, centre, samples))
        if centre / math.sqrt(2.0) < lowest or centre * math.sqrt(2.0) > highest:
            extrapolated.append(float(centre))
    return np.array(bands, dtype=float), extrapolated


# ---- 表を読む --------------------------------------------------------------

def read_table(file_name, sheet=None):
    """「1 行目が周波数、以降の行が材料」の表を読む。→ (周波数, {材料名: 値})

    ・Excel（.xlsx）と CSV のどちらでもよい
    ・**空欄は飛ばす**（もらった表は測っていない周波数が空いている）
    ・周波数の行は「数値が 3 つ以上並ぶ最初の行」で見つける（見出しの位置に依らない）
    """
    rows = _rows(file_name, sheet)
    head = None
    for index, row in enumerate(rows):
        numbers = [k for k, v in enumerate(row) if _number(v) is not None]
        if len(numbers) >= 3:
            head = index
            break
    if head is None:
        raise ValueError(f"周波数の行が見つかりません: {file_name}")

    columns = [k for k, v in enumerate(rows[head]) if _number(v) is not None]
    frequencies = np.array([_number(rows[head][k]) for k in columns], dtype=float)

    materials = {}
    for row in rows[head + 1:]:
        name = next((str(v).strip() for v in row
                     if v is not None and str(v).strip()
                     and _number(v) is None), None)
        if not name:
            continue
        pairs = [(frequencies[i], _number(row[k]))
                 for i, k in enumerate(columns)
                 if k < len(row) and _number(row[k]) is not None]
        if len(pairs) >= 2:
            materials[name] = (np.array([f for f, _ in pairs], dtype=float),
                               np.array([v for _, v in pairs], dtype=float))
    if not materials:
        raise ValueError(f"材料の行が見つかりません: {file_name}")
    return frequencies, materials


def _rows(file_name, sheet=None):
    if os.path.splitext(file_name)[1].lower() in (".xlsx", ".xlsm"):
        import openpyxl
        book = openpyxl.load_workbook(file_name, data_only=True)
        page = book[sheet] if sheet else book[book.sheetnames[0]]
        return [list(row) for row in page.iter_rows(values_only=True)]
    import csv
    for encoding in ("utf-8-sig", "cp932"):
        try:
            with io.open(file_name, encoding=encoding, newline="") as handle:
                return [row for row in csv.reader(handle)]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"読めませんでした: {file_name}")


def _number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def main(argv=None):
    import argparse

    import absorption as ab

    parser = argparse.ArgumentParser(
        description="細かい周波数の吸音率をオクターブバンドに直す")
    parser.add_argument("table", help="Excel か CSV（1 行目が周波数）")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--bands", type=int, default=8, help="6 か 8")
    args = parser.parse_args(argv)

    centres = ab.octave_bands(args.bands)
    _, materials = read_table(args.table, args.sheet)
    print("材料," + ",".join(f"{v:.0f}Hz" for v in centres))
    for name, (frequencies, values) in materials.items():
        bands, extrapolated = to_bands(frequencies, values, centres)
        print(f"{name}," + ",".join(f"{v:.3f}" for v in bands)
              + (f"  ※{[int(v) for v in extrapolated]} Hz は端の値を保持"
                 if extrapolated else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
